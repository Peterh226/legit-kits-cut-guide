"""
Legit Kits — Progress Tracking Workbook Generator
===================================================
Generates a progress-tracking Excel workbook for a Legit Kits quilt project.

Sheets produced:
    1. How To Use       — instructions
    2. Fabric Inventory — verify all fabrics are present and labeled
    3. Cutting Plan     — all pieces grouped by block, sorted simplest first
    4. Block Tracker    — one row per block, track assembly status
    5. Piece Count      — per-fabric piece count vs expected
    6. Final Assembly   — guide for joining completed blocks

Usage:
    python tracking.py --quilt-id skulliver
    python tracking.py --quilt-id land-of-the-free
    python tracking.py --quilt-id skulliver --output MyTracker.xlsx

Requirements:
    pip install openpyxl
"""

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _font(color="000000", bold=False, size=10):
    return Font(name="Arial", color=color, bold=bold, size=size)


def _border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


F_HEADER    = _fill("2F4F4F")
F_SECTION   = _fill("4A7BA7")
F_BLOCK     = _fill("5B8DB8")
F_ALT       = _fill("F2F2F2")
F_WHITE     = _fill("FFFFFF")
F_GREEN     = _fill("C6EFCE")
F_ORANGE    = _fill("FFEB9C")
F_YELLOW    = _fill("FFFFCC")
F_TITLE     = _fill("1F3F6F")

FN_HEADER   = _font("FFFFFF", bold=True, size=11)
FN_SECTION  = _font("FFFFFF", bold=True, size=10)
FN_BLOCK    = _font("FFFFFF", bold=True, size=10)
FN_TITLE    = _font("FFFFFF", bold=True, size=14)
FN_DATA     = _font(size=10)
FN_BOLD     = _font(bold=True, size=10)

THIN_BDR    = _border("thin",   "CCCCCC")
MED_BDR     = _border("medium", "888888")

AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left",   vertical="center", indent=1)
AW = Alignment(horizontal="left",   vertical="top",    wrap_text=True, indent=1)


# ── Low-level cell writers ────────────────────────────────────────────────────

def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.border, c.alignment = FN_HEADER, F_HEADER, THIN_BDR, AC
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _section_row(ws, row, col_start, col_end, title):
    c = ws.cell(row=row, column=col_start, value=title)
    c.font, c.fill, c.border, c.alignment = FN_SECTION, F_SECTION, MED_BDR, AL
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row,   end_column=col_end)
    ws.row_dimensions[row].height = 18
    return c


def _block_row(ws, row, col_start, col_end, title):
    c = ws.cell(row=row, column=col_start, value=title)
    c.font, c.fill, c.border, c.alignment = FN_BLOCK, F_BLOCK, THIN_BDR, AL
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row,   end_column=col_end)
    ws.row_dimensions[row].height = 16
    return c


def _data(ws, row, col, value, center=False, bold=False, status=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = FN_BOLD if bold else FN_DATA
    c.border    = THIN_BDR
    c.alignment = AC if center else AL
    c.fill      = F_YELLOW if status else (F_ALT if row % 2 == 0 else F_WHITE)
    return c


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_quilt(quilt_id):
    root = Path(__file__).parent
    quilt_dir = root / "quilts" / quilt_id
    if not quilt_dir.exists():
        raise SystemExit(f"Error: quilts/{quilt_id}/ not found")
    cg = {}
    exec((quilt_dir / "cut_guide_data.py").read_text(encoding="utf-8"), cg)
    ag = {}
    exec((quilt_dir / "assembly_data.py").read_text(encoding="utf-8"), ag)
    config_path = quilt_dir / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8")) if config_path.exists() else {}
    quilt_name = config.get("quilt_name", quilt_id)
    return cg["DATA"], ag["BLOCKS"], quilt_name


def _default_quilt_id():
    quilts_dir = Path(__file__).parent / "quilts"
    ids = sorted(p.name for p in quilts_dir.iterdir() if p.is_dir()) if quilts_dir.exists() else []
    if not ids:
        raise SystemExit("Error: no quilts found in quilts/")
    return ids[0]


def _output_name(quilt_id, quilt_name):
    slug = "".join(w.capitalize() for w in quilt_name.split())
    return str(Path(__file__).parent / "quilts" / quilt_id / f"{slug}_Tracker.xlsx")


# ── Data index ────────────────────────────────────────────────────────────────

def _build_index(DATA, BLOCKS):
    fabric_info = {}
    by_fragment = defaultdict(list)

    for code, name, sku, size, _piece_num, tmpl, asm_seq, page in DATA:
        if code not in fabric_info:
            fabric_info[code] = {"name": name, "sku": sku, "size": size, "page": page}
        by_fragment[tmpl].append((asm_seq, code, name))

    for tmpl in by_fragment:
        by_fragment[tmpl].sort(key=lambda x: x[0])

    block_stats = {}
    for block_id, frags in BLOCKS.items():
        piece_count = sum(len(by_fragment.get(f, [])) for f in frags)
        block_stats[block_id] = {
            "frag_count":  len(frags),
            "piece_count": piece_count,
            "frags":       frags,
        }

    return fabric_info, by_fragment, block_stats


def _sorted_blocks(block_stats, BLOCKS):
    row_order = "ABCDEFGH"
    def key(bid):
        bs = block_stats[bid]
        return (bs["frag_count"], bs["piece_count"],
                row_order.index(bid[0]), int(bid[1]))
    return sorted(BLOCKS.keys(), key=key)


def _complexity_label(frag_count):
    if frag_count == 1:  return "Simple"
    if frag_count <= 3:  return "Easy"
    if frag_count <= 6:  return "Moderate"
    if frag_count <= 10: return "Complex"
    return "Very Complex"


# ── Sheet 1: How To Use ───────────────────────────────────────────────────────

def build_howto_sheet(wb, quilt_name, fabric_count, single_frag_count):
    ws = wb.active
    ws.title = "How To Use"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 75

    t = ws.cell(row=2, column=2,
                value=f"{quilt_name}  —  Legit Kits Progress Tracker")
    t.font, t.fill, t.alignment = FN_TITLE, F_TITLE, AL
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 36

    rows = [
        (4,  "bold",  "SHEET",            "PURPOSE"),
        (5,  "data",  "Fabric Inventory", f"Confirm all {fabric_count} fabrics are in your kit. "
                                           "Type ✓ in the In Kit and Labeled columns as you verify each one."),
        (6,  "data",  "Cutting Plan",     "Cut your pieces block by block, simplest blocks first. "
                                          "Type ✓ in the Cut column after cutting each piece. "
                                          "Within each fragment, pieces are listed in assembly sequence order — "
                                          "that is the order to sew them onto the pattern paper."),
        (7,  "data",  "Block Tracker",    "Track assembly status for each block. "
                                          "Leave Status blank = Not Started,  ✓ = Complete,  ~ = In Progress. "
                                          "Blocks are sorted from simplest (1 fragment) to most complex."),
        (8,  "data",  "Final Assembly",   "Track the final joining of blocks into the complete quilt top. "
                                          "Follow the numbered steps: Pairs → 4s → 8s → 16s → 32s → Final seam."),
        (10, "tip",   "TIP",              f"Start with the {single_frag_count} single-fragment blocks — they are the simplest "
                                          "and great for building your foundation-paper-piecing confidence."),
        (11, "tip",   "TIP",              "The assembly sequence number (Asm Seq) in the Cutting Plan shows "
                                          "which order to sew each fabric piece onto the fragment pattern paper."),
        (12, "tip",   "TIP",              "The Color Map in your kit booklet is a mirror image of the finished "
                                          "quilt — fabric is sewn to the back of the pattern paper."),
    ]

    for row_num, style, label, text in rows:
        lc = ws.cell(row=row_num, column=2, value=label)
        tc = ws.cell(row=row_num, column=3, value=text)
        tc.alignment = AW
        ws.row_dimensions[row_num].height = 40
        if style == "bold":
            lc.font = tc.font = _font("FFFFFF", bold=True, size=10)
            lc.fill = tc.fill = F_SECTION
        elif style == "tip":
            lc.font = FN_BOLD
            lc.fill = F_ALT
            tc.font = FN_DATA
        else:
            lc.font = FN_BOLD
            lc.fill = F_ALT
            tc.font = FN_DATA
            tc.fill = F_WHITE


# ── Sheet 2: Fabric Inventory ─────────────────────────────────────────────────

def build_inventory_sheet(wb, fabric_info):
    ws = wb.create_sheet("Fabric Inventory")

    cols = [
        ("Code",        8),
        ("Fabric Name", 18),
        ("SKU",          8),
        ("Kit Size",    20),
        ("Page",         8),
        ("✓ In Kit",    9),
        ("✓ Labeled",   9),
    ]
    for col, (label, width) in enumerate(cols, 1):
        _hdr(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    last_row = len(fabric_info) + 1
    for row_idx, (code, info) in enumerate(sorted(fabric_info.items()), 2):
        _data(ws, row_idx, 1, code,         center=True, bold=True)
        _data(ws, row_idx, 2, info["name"])
        _data(ws, row_idx, 3, info["sku"],  center=True)
        _data(ws, row_idx, 4, info["size"])
        _data(ws, row_idx, 5, info["page"], center=True)
        _data(ws, row_idx, 6, "",           center=True, status=True)
        _data(ws, row_idx, 7, "",           center=True, status=True)

    for col_letter in ("F", "G"):
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            CellIsRule(operator="equal", formula=['"✓"'], fill=F_GREEN)
        )
    ws.auto_filter.ref = f"A1:G{last_row}"


# ── Sheet 3: Cutting Plan ─────────────────────────────────────────────────────

def build_cutting_plan_sheet(wb, by_fragment, block_stats, BLOCKS):
    ws = wb.create_sheet("Cutting Plan")

    cols = [
        ("Block",       8),
        ("Fragment",    9),
        ("Asm Seq",     8),
        ("Fabric",      7),
        ("Fabric Name", 16),
        ("✓ Cut",  7),
    ]
    for col, (label, width) in enumerate(cols, 1):
        _hdr(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row = 2
    last_status_row = 2

    for bid in _sorted_blocks(block_stats, BLOCKS):
        stats = block_stats[bid]
        if stats["piece_count"] == 0:
            continue

        frag_label = ("1 fragment" if stats["frag_count"] == 1
                      else f"{stats['frag_count']} fragments")
        title = (f"Block {bid}   —   {frag_label},  "
                 f"{stats['piece_count']} pieces   "
                 f"[{_complexity_label(stats['frag_count'])}]")
        _block_row(ws, row, 1, 6, title)
        row += 1

        for frag_id in stats["frags"]:
            for asm_seq, fab_code, fab_name in by_fragment.get(frag_id, []):
                _data(ws, row, 1, bid,      center=True, bold=True)
                _data(ws, row, 2, frag_id,  center=True)
                _data(ws, row, 3, asm_seq,  center=True)
                _data(ws, row, 4, fab_code, center=True, bold=True)
                _data(ws, row, 5, fab_name)
                _data(ws, row, 6, "",       center=True, status=True)
                last_status_row = row
                row += 1

    ws.conditional_formatting.add(
        f"F2:F{last_status_row}",
        CellIsRule(operator="equal", formula=['"✓"'], fill=F_GREEN)
    )
    ws.auto_filter.ref = f"A1:F{last_status_row}"


# ── Sheet 4: Block Tracker ────────────────────────────────────────────────────

def build_block_tracker_sheet(wb, block_stats, BLOCKS):
    ws = wb.create_sheet("Block Tracker")

    cols = [
        ("Block",       8),
        ("Fragments",  11),
        ("Pieces",      8),
        ("Complexity", 14),
        ("Status",     16),
    ]
    for col, (label, width) in enumerate(cols, 1):
        _hdr(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    sorted_bids = _sorted_blocks(block_stats, BLOCKS)
    last_row = len(sorted_bids) + 1

    for row_idx, bid in enumerate(sorted_bids, 2):
        stats = block_stats[bid]
        _data(ws, row_idx, 1, bid,                                    center=True, bold=True)
        _data(ws, row_idx, 2, stats["frag_count"],                    center=True)
        _data(ws, row_idx, 3, stats["piece_count"],                   center=True)
        _data(ws, row_idx, 4, _complexity_label(stats["frag_count"]), center=True)
        _data(ws, row_idx, 5, "",                                     center=True, status=True)

    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="equal", formula=['"✓"'], fill=F_GREEN)
    )
    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="equal", formula=['"~"'], fill=F_ORANGE)
    )
    ws.auto_filter.ref = f"A1:E{last_row}"


# ── Sheet 5: Piece Count by Fabric ───────────────────────────────────────────

def build_piece_summary_sheet(wb, fabric_info, DATA):
    ws = wb.create_sheet("Piece Count by Fabric")
    legit_total = len(DATA)

    cols = [
        ("Code",         8),
        ("Fabric Name", 18),
        ("In Data",     10),
        ("Expected",    10),
        ("Gap",          7),
        ("Notes",       38),
    ]
    for col, (label, width) in enumerate(cols, 1):
        _hdr(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    piece_counts = defaultdict(int)
    for code, name, sku, size, piece_num, tmpl, asm_seq, page in DATA:
        if piece_num > piece_counts[code]:
            piece_counts[code] = piece_num

    total_in_data = 0
    total_expected = 0

    for row_idx, (code, info) in enumerate(sorted(fabric_info.items()), 2):
        in_data  = piece_counts.get(code, 0)
        expected = in_data
        gap      = expected - in_data
        total_in_data  += in_data
        total_expected += expected

        _data(ws, row_idx, 1, code,     center=True, bold=True)
        _data(ws, row_idx, 2, info["name"])
        _data(ws, row_idx, 3, in_data,  center=True)
        _data(ws, row_idx, 4, expected, center=True)
        c = _data(ws, row_idx, 5, gap if gap else "", center=True)
        if gap > 0:
            c.fill = _fill("FFCCCC")
        _data(ws, row_idx, 6, "")

    total_row = len(fabric_info) + 2
    ws.row_dimensions[total_row].height = 22
    for col in range(1, 7):
        c = ws.cell(row=total_row, column=col)
        c.fill, c.font, c.border, c.alignment = F_HEADER, FN_HEADER, THIN_BDR, AC
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2,
            value=f"Total rows in data: {legit_total}").alignment = AL
    ws.cell(row=total_row, column=3, value=total_in_data)
    ws.cell(row=total_row, column=4, value=total_expected)
    ws.cell(row=total_row, column=5, value=total_expected - total_in_data)

    ws.auto_filter.ref = f"A1:F{total_row - 1}"


# ── Sheet 6: Final Assembly ───────────────────────────────────────────────────

def build_final_assembly_sheet(wb, quilt_name, block_count):
    ws = wb.create_sheet("Final Assembly")
    ws.sheet_view.showGridLines = False

    _section_row(ws, 1, 1, 4,
                 f"Final Quilt Top Assembly  —  {quilt_name}  ({block_count} blocks)")
    ws.row_dimensions[1].height = 22

    cols = [("Step", 6), ("What to Join", 55), ("# Groups", 10), ("✓ Done", 8)]
    for col, (label, width) in enumerate(cols, 1):
        _hdr(ws, 2, col, label, width)

    steps = [
        (1, "Join adjacent column pairs within each row  "
            "(columns 1+2, 3+4, 5+6, 7+8  ×  8 rows = 32 pairs)", 32),
        (2, "Join pairs into 2×2 groups of 4 blocks", 16),
        (3, "Join groups of 4 into 4×2 groups of 8 blocks", 8),
        (4, "Join groups of 8 into 4×4 groups of 16 blocks", 4),
        (5, "Join groups of 16 into two halves of 32 blocks each  "
            "(leaves one final horizontal seam)", 2),
        (6, "Sew the final horizontal seam  —  quilt top is complete!", 1),
    ]

    for row_idx, (step, desc, groups) in enumerate(steps, 3):
        _data(ws, row_idx, 1, step,   center=True, bold=True)
        _data(ws, row_idx, 2, desc)
        _data(ws, row_idx, 3, groups, center=True)
        _data(ws, row_idx, 4, "",     center=True, status=True)
        ws.row_dimensions[row_idx].height = 22

    ws.conditional_formatting.add(
        "D3:D8",
        CellIsRule(operator="equal", formula=['"✓"'], fill=F_GREEN)
    )


# ── Entry point ───────────────────────────────────────────────────────────────

def generate(quilt_id, output_path=None):
    DATA, BLOCKS, quilt_name = _load_quilt(quilt_id)
    if output_path is None:
        output_path = _output_name(quilt_id, quilt_name)

    fabric_info, by_fragment, block_stats = _build_index(DATA, BLOCKS)
    single_frag_count = sum(1 for bs in block_stats.values() if bs["frag_count"] == 1)
    block_count = len(BLOCKS)

    wb = Workbook()
    build_howto_sheet(wb, quilt_name, len(fabric_info), single_frag_count)
    build_inventory_sheet(wb, fabric_info)
    build_cutting_plan_sheet(wb, by_fragment, block_stats, BLOCKS)
    build_block_tracker_sheet(wb, block_stats, BLOCKS)
    build_piece_summary_sheet(wb, fabric_info, DATA)
    build_final_assembly_sheet(wb, quilt_name, block_count)

    wb.save(output_path)

    total_pieces = sum(bs["piece_count"] for bs in block_stats.values())
    multi_frag = block_count - single_frag_count

    print(f"Generated  : {output_path}")
    print(f"  Quilt    : {quilt_name}")
    print(f"  Fabrics  : {len(fabric_info)}")
    print(f"  Blocks   : {block_count}  ({single_frag_count} simple, {multi_frag} multi-fragment)")
    print(f"  Pieces   : {total_pieces}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate Legit Kits progress-tracking spreadsheet"
    )
    parser.add_argument("--quilt-id", "-q", help="Quilt ID (default: first in quilts/)")
    parser.add_argument("--output",   "-o", help="Output filename (default: <QuiltName>_Tracker.xlsx)")
    args = parser.parse_args()
    generate(args.quilt_id or _default_quilt_id(), args.output)
