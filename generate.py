"""
Legit Kits Cut Guide Generator
================================
Reads fabric and piece data from quilts/<quilt-id>/cut_guide_data.py and produces
a formatted Excel workbook: <QuiltName>_CutGuide.xlsx

Usage:
    python generate.py --quilt-id skulliver
    python generate.py --quilt-id land-of-the-free
    python generate.py --quilt-id skulliver --output my_custom_name.xlsx

Requirements:
    pip install openpyxl
"""

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL     = PatternFill("solid", start_color="2F4F4F")
STAT_FILL       = PatternFill("solid", start_color="1F3F6F")
SECTION_FILL    = PatternFill("solid", start_color="4A7BA7")
ALT_FILL        = PatternFill("solid", start_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
STAT_LABEL_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
STAT_VALUE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=20)
SECTION_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT       = Font(name="Arial", size=10)
BOLD_FONT       = Font(name="Arial", bold=True, size=10)
THIN            = Side(style="thin", color="CCCCCC")
BORDER          = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED             = Side(style="medium", color="888888")
MED_BORDER      = Border(left=MED, right=MED, top=MED, bottom=MED)


def _header_cell(ws, row, col, value, width):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _data_cell(ws, row, col, value, center=False, bold=False):
    fill = ALT_FILL if row % 2 == 0 else WHITE_FILL
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = BOLD_FONT if bold else DATA_FONT
    cell.fill      = fill
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    return cell


def _stat_block(ws, row, col, label, value, label_width=None):
    lc = ws.cell(row=row,     column=col, value=label)
    vc = ws.cell(row=row + 1, column=col, value=value)
    lc.font      = STAT_LABEL_FONT
    lc.fill      = STAT_FILL
    lc.border    = MED_BORDER
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc.font      = STAT_VALUE_FONT
    vc.fill      = STAT_FILL
    vc.border    = MED_BORDER
    vc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height     = 22
    ws.row_dimensions[row + 1].height = 42
    if label_width:
        ws.column_dimensions[get_column_letter(col)].width = label_width


def _section_header(ws, row, col_start, col_end, title):
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font      = SECTION_FONT
    cell.fill      = SECTION_FILL
    cell.border    = MED_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    ws.row_dimensions[row].height = 20


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_quilt(quilt_id):
    root = Path(__file__).parent
    quilt_dir = root / "quilts" / quilt_id
    if not quilt_dir.exists():
        raise SystemExit(f"Error: quilts/{quilt_id}/ not found")
    g = {}
    exec((quilt_dir / "cut_guide_data.py").read_text(encoding="utf-8"), g)
    config_path = quilt_dir / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8")) if config_path.exists() else {}
    quilt_name = config.get("quilt_name", quilt_id)
    return g["DATA"], quilt_name


def _default_quilt_id():
    quilts_dir = Path(__file__).parent / "quilts"
    ids = sorted(p.name for p in quilts_dir.iterdir() if p.is_dir()) if quilts_dir.exists() else []
    if not ids:
        raise SystemExit("Error: no quilts found in quilts/")
    return ids[0]


def _output_name(quilt_id, quilt_name):
    slug = "".join(w.capitalize() for w in quilt_name.split())
    return str(Path(__file__).parent / "quilts" / quilt_id / f"{slug}_CutGuide.xlsx")


# ---------------------------------------------------------------------------
# Aggregate statistics from DATA
# ---------------------------------------------------------------------------

def _compute_stats(DATA):
    fabric_info   = {}
    template_freq = defaultdict(int)
    page_fabrics  = defaultdict(list)
    size_counts   = defaultdict(int)
    size_cuts     = defaultdict(int)

    for row in DATA:
        code, name, sku, size, piece, template, qty, page = row

        if code not in fabric_info:
            fabric_info[code] = {
                "name": name, "sku": sku, "size": size,
                "page": page, "piece_count": 0, "total_cuts": 0
            }
            page_fabrics[page].append(code)

        fabric_info[code]["piece_count"] += 1
        fabric_info[code]["total_cuts"]  += qty
        template_freq[template]          += qty
        size_counts[size]                += 1
        size_cuts[size]                  += qty

    two_block_pages = {pg: codes for pg, codes in page_fabrics.items() if len(codes) == 2}

    return {
        "fabric_info":     fabric_info,
        "total_cuts":      sum(r[6] for r in DATA),
        "total_pieces":    len(DATA),
        "total_fabrics":   len(fabric_info),
        "top_templates":   sorted(template_freq.items(), key=lambda x: -x[1])[:15],
        "most_pieces":     sorted(fabric_info.items(), key=lambda x: -x[1]["piece_count"])[:10],
        "most_cuts":       sorted(fabric_info.items(), key=lambda x: -x[1]["total_cuts"])[:10],
        "size_counts":     size_counts,
        "size_cuts":       size_cuts,
        "two_block_pages": two_block_pages,
        "page_fabrics":    page_fabrics,
        "max_page":        max(r[7] for r in DATA),
    }


# ---------------------------------------------------------------------------
# Sheet: Cut Guide
# ---------------------------------------------------------------------------

def build_cut_guide_sheet(wb, DATA):
    ws = wb.active
    ws.title = "Cut Guide"

    headers = [
        ("Fabric Code",   14),
        ("Fabric Name",   16),
        ("SKU",           10),
        ("Fabric Size",   22),
        ("Cut #",         10),
        ("Segment ID",    16),
        ("Sew Sequence",  12),
        ("Page",          10),
    ]
    for col, (label, width) in enumerate(headers, 1):
        _header_cell(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28

    for row_idx, (code, name, sku, size, piece, template, qty, page) in enumerate(DATA, 2):
        _data_cell(ws, row_idx, 1, code,     center=True)
        _data_cell(ws, row_idx, 2, name)
        _data_cell(ws, row_idx, 3, sku)
        _data_cell(ws, row_idx, 4, size)
        _data_cell(ws, row_idx, 5, piece,    center=True)
        _data_cell(ws, row_idx, 6, template, center=True)
        _data_cell(ws, row_idx, 7, qty,      center=True)
        _data_cell(ws, row_idx, 8, page,     center=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(DATA) + 1}"


# ---------------------------------------------------------------------------
# Sheet: By Fabric Code
# ---------------------------------------------------------------------------

def build_summary_sheet(wb, stats):
    ws = wb.create_sheet("By Fabric Code")

    headers = [
        ("Fabric Code",  14),
        ("Fabric Name",  18),
        ("SKU",          10),
        ("Fabric Size",  22),
        ("Page",         10),
        ("Piece Count",  12),
        ("Total Cuts",   12),
    ]
    for col, (label, width) in enumerate(headers, 1):
        _header_cell(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28

    for row_idx, (code, info) in enumerate(sorted(stats["fabric_info"].items()), 2):
        _data_cell(ws, row_idx, 1, code,               center=True)
        _data_cell(ws, row_idx, 2, info["name"])
        _data_cell(ws, row_idx, 3, info["sku"],         center=True)
        _data_cell(ws, row_idx, 4, info["size"])
        _data_cell(ws, row_idx, 5, info["page"],        center=True)
        _data_cell(ws, row_idx, 6, info["piece_count"], center=True)
        _data_cell(ws, row_idx, 7, info["total_cuts"],  center=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(stats['fabric_info']) + 1}"


# ---------------------------------------------------------------------------
# Sheet: By Page
# ---------------------------------------------------------------------------

def build_page_sheet(wb, stats):
    ws = wb.create_sheet("By Page")

    headers = [
        ("Page",         8),
        ("Fabric Code", 14),
        ("Fabric Name", 18),
        ("Total Pieces", 14),
    ]
    for col, (label, width) in enumerate(headers, 1):
        _header_cell(ws, 1, col, label, width)
    ws.row_dimensions[1].height = 28

    page_rows = defaultdict(list)
    for code, info in stats["fabric_info"].items():
        page_rows[info["page"]].append((code, info["name"], info["piece_count"]))

    row_idx = 2
    for page in sorted(page_rows.keys()):
        for code, name, piece_count in sorted(page_rows[page]):
            _data_cell(ws, row_idx, 1, page,        center=True)
            _data_cell(ws, row_idx, 2, code,        center=True)
            _data_cell(ws, row_idx, 3, name)
            _data_cell(ws, row_idx, 4, piece_count, center=True)
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{row_idx - 1}"


# ---------------------------------------------------------------------------
# Sheet: Statistics
# ---------------------------------------------------------------------------

def build_stats_sheet(wb, stats):
    ws = wb.create_sheet("Statistics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    _stat_block(ws, 2, 2, "Total Fabrics",  stats["total_fabrics"],        label_width=18)
    _stat_block(ws, 2, 3, "Piece Rows",     stats["total_pieces"],         label_width=18)
    _stat_block(ws, 2, 4, "Total Cuts",     stats["total_cuts"],           label_width=18)
    _stat_block(ws, 2, 5, "Pages",          stats["max_page"],             label_width=18)
    _stat_block(ws, 2, 6, "2-Block Pages",  len(stats["two_block_pages"]), label_width=18)

    row = 6

    _section_header(ws, row, 2, 4, "Top 15 Templates by Total Cuts")
    row += 1
    _header_cell(ws, row, 2, "Template",      14)
    _header_cell(ws, row, 3, "Total Cuts",    14)
    _header_cell(ws, row, 4, "% of All Cuts", 16)
    row += 1
    for tmpl, cuts in stats["top_templates"]:
        pct = cuts / stats["total_cuts"] * 100
        _data_cell(ws, row, 2, tmpl,          center=True)
        _data_cell(ws, row, 3, cuts,          center=True)
        _data_cell(ws, row, 4, f"{pct:.1f}%", center=True)
        row += 1
    row += 1

    _section_header(ws, row, 2, 5, "Top 10 Fabrics by Piece Rows")
    row += 1
    _header_cell(ws, row, 2, "Code",        10)
    _header_cell(ws, row, 3, "Fabric Name", 18)
    _header_cell(ws, row, 4, "Piece Rows",  12)
    _header_cell(ws, row, 5, "Total Cuts",  12)
    row += 1
    for code, info in stats["most_pieces"]:
        _data_cell(ws, row, 2, code,               center=True, bold=True)
        _data_cell(ws, row, 3, info["name"])
        _data_cell(ws, row, 4, info["piece_count"], center=True)
        _data_cell(ws, row, 5, info["total_cuts"],  center=True)
        row += 1
    row += 1

    _section_header(ws, row, 2, 5, "Top 10 Fabrics by Total Cuts")
    row += 1
    _header_cell(ws, row, 2, "Code",        10)
    _header_cell(ws, row, 3, "Fabric Name", 18)
    _header_cell(ws, row, 4, "Piece Rows",  12)
    _header_cell(ws, row, 5, "Total Cuts",  12)
    row += 1
    for code, info in stats["most_cuts"]:
        _data_cell(ws, row, 2, code,               center=True, bold=True)
        _data_cell(ws, row, 3, info["name"])
        _data_cell(ws, row, 4, info["piece_count"], center=True)
        _data_cell(ws, row, 5, info["total_cuts"],  center=True)
        row += 1
    row += 1

    _section_header(ws, row, 2, 5, "Fabric Sizes — Fabrics per Size Category")
    row += 1
    _header_cell(ws, row, 2, "Fabric Size",      22)
    _header_cell(ws, row, 3, "# Fabrics",        12)
    _header_cell(ws, row, 4, "Total Piece Rows", 16)
    _header_cell(ws, row, 5, "Total Cuts",       12)
    row += 1
    for size, count in sorted(stats["size_counts"].items(), key=lambda x: -x[1]):
        _data_cell(ws, row, 2, size)
        _data_cell(ws, row, 3, count,                      center=True)
        _data_cell(ws, row, 4, stats["size_counts"][size], center=True)
        _data_cell(ws, row, 5, stats["size_cuts"][size],   center=True)
        row += 1
    row += 1

    _section_header(ws, row, 2, 5,
                    f"Pages with Two Fabrics ({len(stats['two_block_pages'])} pages)")
    row += 1
    _header_cell(ws, row, 2, "Page",       10)
    _header_cell(ws, row, 3, "Fabric 1",   14)
    _header_cell(ws, row, 4, "Fabric 2",   14)
    _header_cell(ws, row, 5, "Fabric Names", 32)
    row += 1
    for pg in sorted(stats["two_block_pages"].keys()):
        codes = stats["two_block_pages"][pg]
        names = " / ".join(stats["fabric_info"][c]["name"] for c in codes)
        _data_cell(ws, row, 2, pg,       center=True)
        _data_cell(ws, row, 3, codes[0], center=True, bold=True)
        _data_cell(ws, row, 4, codes[1], center=True, bold=True)
        _data_cell(ws, row, 5, names)
        row += 1


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def generate(quilt_id, output_path=None):
    DATA, quilt_name = _load_quilt(quilt_id)
    if output_path is None:
        output_path = _output_name(quilt_id, quilt_name)
    stats = _compute_stats(DATA)
    wb = Workbook()
    build_cut_guide_sheet(wb, DATA)
    build_summary_sheet(wb, stats)
    build_page_sheet(wb, stats)
    build_stats_sheet(wb, stats)
    wb.save(output_path)

    print(f"Generated : {output_path}")
    print(f"  Quilt   : {quilt_name}")
    print(f"  Fabrics : {stats['total_fabrics']}")
    print(f"  Pieces  : {stats['total_pieces']}")
    print(f"  Cuts    : {stats['total_cuts']}")
    print(f"  2-block pages: {len(stats['two_block_pages'])}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Legit Kits Cut Guide spreadsheet")
    parser.add_argument("--quilt-id", "-q", help="Quilt ID (default: first in quilts/)")
    parser.add_argument("--output",   "-o", help="Output filename (default: <QuiltName>_CutGuide.xlsx)")
    args = parser.parse_args()
    generate(args.quilt_id or _default_quilt_id(), args.output)
